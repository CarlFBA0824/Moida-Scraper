"""
Master matcher/merger: Moida scrape + ScanUnlimited export + Keepa export.

Copied from the sibling Web-Scraping repo's master.py (kept in sync there
too) so it can run directly from this repo without switching folders -
purely a convenience copy, not a fork with different logic.

Run this after each scrape.py run, once you've pushed the resulting GTINs
through ScanUnlimited and Keepa and downloaded their exports. It matches all
three files by GTIN/UPC/EAN (normalizing leading zeros), recomputes
Cost/Profit/ROI/Margin using your freshest scraped price as the true
cost basis, and flags rows worth a second look.

Setup:
    pip install openpyxl

Run:
    python master.py --vitacost output/moidaus_kundal_20260831_104821.csv \
                      --scanunlimited ScanUnlimited_export.csv \
                      --keepa KeepaExport.csv \
                      --source-name Moida \
                      --output output/master_kundal.xlsx

If you omit --vitacost, it auto-picks the newest output/moidaus_*.csv.

Match key: GTIN/UPC/EAN, normalized by stripping leading zeros (so our
12-digit UPC like 808124111806 matches ScanUnlimited's 0808124111806 EAN and
Keepa's 00808124111806 GTIN). Keepa's UPC/EAN/GTIN columns sometimes hold
multiple comma-separated codes; all are indexed.

Price logic:
    cost        = our scraped price (Moida's cart_price, the freshest
                  post-discount number - always the true cost basis)
    sell_price  = first available of, in order: Keepa "Buy Box: Current"
                  (a live buy box - the reliable number), Keepa
                  "Amazon: Current", ScanUnlimited "Price", Keepa
                  "New: Current" (lowest 3rd-party new offer), ScanUnlimited
                  "Amazon Lowest Price", Keepa "List Price: Current" (MSRP -
                  the weakest fallback). sell_price_source records which one
                  was actually used, so you can judge how trustworthy the
                  number is at a glance.
    referral_fee = Keepa "Referral Fee based on current Buy Box price",
                   else ScanUnlimited "Referral Fee"
    fba_fee     = Keepa "FBA Pick&Pack Fee", else ScanUnlimited "Fba Fee"
    profit      = sell_price - cost - referral_fee - fba_fee
    roi_pct     = profit / cost * 100
    margin_pct  = profit / sell_price * 100

This mirrors the formula ScanUnlimited itself uses (verified against its own
Profit/ROI columns), just recomputed with your latest scraped cost and
Keepa's fresher fee/price data where available.

A missing buy box no longer blanks out Profit/ROI/Margin - it just means
sell_price came from one of the weaker fallbacks above instead of a live
buy box. The row still gets matched, priced, and ranked; the no_buy_box
flag (plus the sell_price_source column) is your signal to sanity-check
the number yourself before treating it as a real opportunity.

Flags:
    no_amazon_match          - not found in ScanUnlimited or Keepa at all
    price_changed_since_scan - ScanUnlimited's Cost differs from the current
                                scraped price by more than $0.01
    no_buy_box                - no live Buy Box price on Amazon right now;
                                sell_price (if any) came from a weaker
                                fallback (Amazon: Current, New offer,
                                ScanUnlimited price, or List Price) - still
                                populated for review, just not a confirmed
                                live buy box price
    no_referral_fee_data      - referral fee couldn't be determined from
                                either source (it's normally tied to a live
                                buy box price, so this fires together with
                                no_buy_box), and Profit/ROI/Margin above had
                                to assume $0 referral fee to be computable
                                at all - the real number is commonly 8-17%
                                of sell price lower than shown, so treat
                                these numbers as an upper bound, not a
                                reliable estimate
    contents_count_mismatch  - the item count inside one unit, per the
                                scraped product name (e.g. "180 Capsules"),
                                doesn't match Amazon's Size field count (e.g.
                                "180 Count (Pack of 1)"). Best-effort text
                                extraction, always spot-check flagged rows.
    bundle_mismatch           - Amazon's matched listing bundles multiple
                                retail units together (Keepa/ScanUnlimited's
                                "Package Quantity" > 1), while the scrape
                                source sells one unit per SKU. E.g. one box
                                of crackers, but the matched ASIN is a
                                "Pack of 6" of that same box - the prices
                                aren't comparable as-is.
    ambiguous_match            - this GTIN/UPC is shared by more than one
                                distinct ASIN in Keepa's data (common with
                                reused barcodes across resellers/duplicate
                                listings). The best candidate is auto-picked
                                (prefers an active Buy Box price, then more
                                reviews), but always verify manually - the
                                picked ASIN may not be the intended product.

A GTIN can trigger more than one flag at once (e.g. an ambiguous match that
also turns out to be a bundle mismatch) - treat any flagged row as "needs a
manual look," not as a specific single problem.
"""

import argparse
import csv
import glob
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def normalize_code(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    if s.isdigit():
        return str(int(s))
    return s


def normalize_codes_field(raw: Any) -> List[str]:
    if raw is None:
        return []
    parts = re.split(r"[;,]", str(raw))
    return [c for c in (normalize_code(p) for p in parts) if c]


def to_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    s = str(x).strip().replace("%", "").replace(",", "").replace("$", "")
    if s in ("", "-", "N/A", "n/a"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def su_float(row: Optional[Dict[str, Any]], *names: str) -> Optional[float]:
    """to_float for a ScanUnlimited field, but treats a literal 0 as 'no
    data' too. ScanUnlimited writes Price/Referral Fee/Fba Fee as exactly 0
    (not blank) for an otherwise-matched ASIN when it has no live price data
    (confirmed: always all three together) - a real Amazon listing is never
    actually priced, referral-fee'd, or FBA-fee'd at $0.00, so treating that
    0 as real data would silently understate cost and inflate ROI."""
    if not row:
        return None
    value = to_float(get_ci(row, *names))
    return value if value else None


def get_ci(row: Dict[str, Any], *names: str) -> Optional[str]:
    """Case-insensitive, whitespace-tolerant column lookup."""
    lower_map = {k.strip().lower(): v for k, v in row.items() if k}
    for name in names:
        v = lower_map.get(name.strip().lower())
        if v not in (None, ""):
            return v
    return None


# Count-style unit words that indicate "N of these" (as opposed to weight/volume
# units like "oz" or "ml", which don't imply a pack count of separate items).
PACK_UNIT_WORDS = (
    r"shakes?|sticks?|pieces?|capsules?|tablets?|gummies|gummy|count|bars?|"
    r"pouches?|cups?|kits?|packs?|bottles?|snacks?|sachets?|servings?|"
    r"disks?|softgels?|strips?|gelcaps?|cartridges?|razors?|patches?|pads?|"
    r"tampons?|toothbrush(?:es)?|candles?|toys?"
)
PACK_QTY_PATTERN = re.compile(rf"\b(\d+)\s*(?:{PACK_UNIT_WORDS})\b", re.IGNORECASE)


def extract_pack_qty(product_name: Optional[str]) -> Optional[int]:
    """Best-effort extraction of a pack/count size from a scraped product
    name, e.g. '8 Shakes' -> 8, '24 Sticks' -> 24. Returns None if no
    count-style unit is found (e.g. the product is a single item, or only
    weight/volume units like 'oz'/'ml' are present)."""
    if not product_name:
        return None
    matches = PACK_QTY_PATTERN.findall(product_name)
    # Product names conventionally put the actual pack count last (e.g.
    # "..., Omega-3 Gummies, 80 Gummies" - "Omega-3" is a red herring).
    return int(matches[-1]) if matches else None


AMAZON_COUNT_PATTERN = re.compile(r"\b(\d+)\s*Count\b", re.IGNORECASE)


def extract_amazon_count(size_field: Optional[str]) -> Optional[int]:
    """Extract the inside-the-package item count from Keepa/ScanUnlimited's
    'Size' field, e.g. '180 Count (Pack of 1)' -> 180. Deliberately only
    matches an explicit 'Count' unit (not oz/ml/etc.) and is NOT the same as
    the 'Package Quantity' field, which means how many outer boxes were
    shipped (almost always 1) rather than how many items are inside one."""
    if not size_field:
        return None
    match = AMAZON_COUNT_PATTERN.search(str(size_field))
    return int(match.group(1)) if match else None


TITLE_PACK_PATTERN = re.compile(r"\bpack of (\d+)\b", re.IGNORECASE)


def extract_title_pack_qty(title: Optional[str]) -> Optional[int]:
    """Extract a 'Pack of N' bundle count from an Amazon title, e.g.
    'Suzie's Spelt Flatbread| Pack of 6 |...' -> 6."""
    if not title:
        return None
    match = TITLE_PACK_PATTERN.search(str(title))
    return int(match.group(1)) if match else None


def read_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def find_latest_vitacost_csv() -> Optional[Path]:
    matches = sorted(glob.glob("output/moidaus_*.csv"))
    return Path(matches[-1]) if matches else None


# --------------------------------------------------------------------------
# Indexing
# --------------------------------------------------------------------------

def index_by_code(rows: List[Dict[str, Any]], *code_fields: str) -> Dict[str, Dict[str, Any]]:
    index: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        for field in code_fields:
            value = get_ci(row, field)
            for code in normalize_codes_field(value):
                index.setdefault(code, row)
    return index


def index_by_code_multi(rows: List[Dict[str, Any]], *code_fields: str) -> Dict[str, List[Dict[str, Any]]]:
    """Like index_by_code, but keeps every distinct row seen for a code
    instead of just the first. Used for Keepa, where the same UPC/EAN/GTIN
    can legitimately be listed under several different ASINs (resellers,
    duplicate listings, or genuinely different pack sizes reusing the same
    barcode)."""
    index: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        for field in code_fields:
            value = get_ci(row, field)
            for code in normalize_codes_field(value):
                index.setdefault(code, []).append(row)
    return index


def pick_best_keepa_match(candidates: List[Dict[str, Any]]) -> tuple:
    """Given every Keepa row sharing a GTIN, dedupe by ASIN and pick the
    most likely 'real' listing: prefer one with an active Buy Box price
    (a listing with no live price can't be sold from anyway), then the
    one with the most reviews (a proxy for being the established/parent
    listing rather than a spinoff duplicate). Returns (best_row, distinct_asin_count)."""
    by_asin: Dict[str, Dict[str, Any]] = {}
    for row in candidates:
        asin = get_ci(row, "ASIN")
        if asin and asin not in by_asin:
            by_asin[asin] = row
    unique = list(by_asin.values())
    if not unique:
        return None, 0
    if len(unique) == 1:
        return unique[0], 1

    def score(row: Dict[str, Any]) -> tuple:
        has_price = to_float(get_ci(row, "Buy Box: Current")) is not None
        reviews = to_float(get_ci(row, "Reviews: Rating Count")) or 0.0
        return (has_price, reviews)

    best = max(unique, key=score)
    return best, len(unique)


# --------------------------------------------------------------------------
# Merge
# --------------------------------------------------------------------------

def build_master_rows(
    vitacost_rows: List[Dict[str, Any]],
    scanunlimited_index: Dict[str, Dict[str, Any]],
    keepa_index: Dict[str, List[Dict[str, Any]]],
) -> List[Dict[str, Any]]:
    master_rows = []

    for v in vitacost_rows:
        code = normalize_code(v.get("gtin"))
        su = scanunlimited_index.get(code) if code else None
        if su and not get_ci(su, "ASIN"):
            # ScanUnlimited emits a placeholder row (blank ASIN/Title,
            # Package Quantity "0") for every submitted GTIN it couldn't
            # find on Amazon at all - treat that the same as no match.
            su = None
        keepa_candidates = keepa_index.get(code, []) if code else []
        keepa_by_asin: Dict[str, Dict[str, Any]] = {}
        for row in keepa_candidates:
            row_asin = get_ci(row, "ASIN")
            if row_asin and row_asin not in keepa_by_asin:
                keepa_by_asin[row_asin] = row
        distinct_asin_count = len(keepa_by_asin)

        su_asin = get_ci(su, "ASIN") if su else None
        if su_asin and su_asin in keepa_by_asin:
            # ScanUnlimited already resolved this GTIN to one specific ASIN -
            # use Keepa's row for that SAME ASIN, never a different one, so
            # every field describes one single real listing.
            keepa = keepa_by_asin[su_asin]
        elif su_asin:
            # ScanUnlimited found an ASIN Keepa doesn't have data for at all -
            # don't guess at a different Keepa listing for this row.
            keepa = None
        else:
            keepa, _ = pick_best_keepa_match(keepa_candidates)

        # Vitacost's scrape CSV has a single "price" column; Moida's has
        # original_price/sale_price/cart_price instead, with cart_price (the
        # welcome30-discounted checkout price) as the real cost basis - the
        # same number submitted to ScanUnlimited for Moida rows. The two
        # column sets are mutually exclusive per source file, so this falls
        # back cleanly without misreading either one.
        cost = to_float(v.get("price") or v.get("cart_price"))

        # Fallback chain, strongest signal first. has_buy_box tracks whether
        # we got a real live buy box price specifically (the only case that
        # doesn't get the no_buy_box flag below) - everything past that is
        # progressively weaker evidence of what the item would actually sell
        # for, kept so the row still gets a Profit/ROI/Margin estimate
        # instead of being left blank.
        sell_price = None
        sell_price_source = None
        has_buy_box = False
        if keepa:
            sell_price = to_float(get_ci(keepa, "Buy Box: Current"))
            if sell_price:
                sell_price_source = "Buy Box (Keepa)"
                has_buy_box = True
        if not sell_price and keepa:
            sell_price = to_float(get_ci(keepa, "Amazon: Current"))
            if sell_price:
                sell_price_source = "Amazon (Keepa)"
        if not sell_price and su:
            sell_price = su_float(su, "Price")
            if sell_price:
                sell_price_source = "Price (ScanUnlimited)"
        if not sell_price and keepa:
            sell_price = to_float(get_ci(keepa, "New: Current"))
            if sell_price:
                sell_price_source = "Lowest New Offer (Keepa)"
        if not sell_price and su:
            sell_price = su_float(su, "Amazon Lowest Price")
            if sell_price:
                sell_price_source = "Amazon Lowest Price (ScanUnlimited)"
        if not sell_price and keepa:
            sell_price = to_float(get_ci(keepa, "List Price: Current"))
            if sell_price:
                sell_price_source = "List Price (Keepa)"

        referral_fee = to_float(get_ci(keepa, "Referral Fee based on current Buy Box price")) if keepa else None
        if referral_fee is None and su:
            referral_fee = su_float(su, "Referral Fee")

        fba_fee = to_float(get_ci(keepa, "FBA Pick&Pack Fee")) if keepa else None
        if fba_fee is None and su:
            fba_fee = su_float(su, "Fba Fee")

        # Neither source has real referral-fee data for this row (it's tied
        # to a live buy box price, which this row doesn't have) - the profit
        # calc below still has to assume $0 so it can produce a number at
        # all, but that makes Profit/ROI/Margin optimistic by whatever the
        # real referral fee would have been (commonly 8-17% of sell price).
        # Flagged below so that's visible instead of silently baked in.
        missing_referral_fee = sell_price is not None and referral_fee is None

        profit = roi_pct = margin_pct = None
        if cost is not None and sell_price:
            rf = referral_fee or 0.0
            ff = fba_fee or 0.0
            profit = round(sell_price - cost - rf - ff, 2)
            roi_pct = round(profit / cost * 100, 1) if cost else None
            margin_pct = round(profit / sell_price * 100, 1) if sell_price else None

        # "Contents count" - items inside one retail unit, e.g. "180 Capsules"
        # inside one bottle. Scraped product name vs. Amazon's Size
        # field ("180 Count (Pack of 1)").
        vitacost_contents_qty = extract_pack_qty(v.get("product_name"))
        amazon_contents_qty = None
        if keepa:
            amazon_contents_qty = extract_amazon_count(get_ci(keepa, "Size"))
        if amazon_contents_qty is None and su:
            amazon_contents_qty = extract_amazon_count(get_ci(su, "Size"))

        # "Bundle count" - how many separate retail units are bundled into
        # ONE Amazon listing, e.g. one box scraped but the matched
        # ASIN is a "Pack of 6" of that same box. The scrape source always
        # sells one bundle per SKU (implicitly 1) unless its own name says
        # otherwise; Amazon's bundle count comes from Keepa/ScanUnlimited's
        # own "Package Quantity" field, or a "Pack of N" mention in the title
        # when that structured field is missing or self-inconsistent (Keepa
        # sometimes has a title saying "Pack of 6" while its own Package
        # Quantity field says 1 for the same ASIN - a source data error we
        # can't resolve, so we take whichever signal indicates a bundle).
        vitacost_bundle_qty = 1
        amazon_title = get_ci(keepa, "Title") if keepa else (get_ci(su, "Title") if su else None)
        amazon_bundle_qty_field = to_float(get_ci(keepa, "Package: Quantity")) if keepa else None
        if amazon_bundle_qty_field is None and su:
            amazon_bundle_qty_field = to_float(get_ci(su, "Package Quantity"))
        if amazon_bundle_qty_field is not None and amazon_bundle_qty_field < 1:
            # A literal 0 is a data gap, not a real "zero-count bundle".
            amazon_bundle_qty_field = None
        amazon_bundle_qty_title = extract_title_pack_qty(amazon_title)
        candidates = [q for q in (amazon_bundle_qty_field, amazon_bundle_qty_title) if q is not None]
        amazon_bundle_qty = max(candidates) if candidates else None

        flags = []
        if not su and not keepa:
            flags.append("no_amazon_match")
        if missing_referral_fee:
            flags.append("no_referral_fee_data (Profit/ROI assumes $0 referral fee)")
        if su:
            su_cost = to_float(get_ci(su, "Cost"))
            if su_cost is not None and cost is not None and abs(su_cost - cost) > 0.01:
                flags.append("price_changed_since_scan")
        if not has_buy_box:
            flags.append("no_buy_box")
        if (
            vitacost_contents_qty is not None
            and amazon_contents_qty is not None
            and int(vitacost_contents_qty) != int(amazon_contents_qty)
        ):
            flags.append(
                f"contents_count_mismatch (vitacost x{int(vitacost_contents_qty)} vs amazon x{int(amazon_contents_qty)})"
            )
        # "Pack of N" in a title is ambiguous: it can mean N separate retail
        # units bundled together (a real mismatch), or just be restating the
        # contents count the scraped name already gives (e.g. both say
        # "8 Shakes"/"Pack of 8" for the same single box) - not a mismatch.
        # Suppress when the two numbers agree, since that's corroboration.
        same_as_contents_count = (
            vitacost_contents_qty is not None
            and amazon_bundle_qty is not None
            and int(vitacost_contents_qty) == int(amazon_bundle_qty)
        )
        if amazon_bundle_qty is not None and int(amazon_bundle_qty) != vitacost_bundle_qty and not same_as_contents_count:
            flags.append(f"bundle_mismatch (vitacost sells 1, amazon listing bundles x{int(amazon_bundle_qty)})")
        if distinct_asin_count > 1:
            flags.append(f"ambiguous_match ({distinct_asin_count} different ASINs share this GTIN)")

        asin = get_ci(su, "ASIN") if su else (get_ci(keepa, "ASIN") if keepa else None)
        brand = get_ci(keepa, "Brand") if keepa else (get_ci(su, "Brand") if su else None)
        title_amazon = amazon_title
        sales_rank = to_float(get_ci(keepa, "Sales Rank: Current")) if keepa else to_float(get_ci(su, "Sales Rank")) if su else None
        review_rating = to_float(get_ci(keepa, "Reviews: Rating")) if keepa else to_float(get_ci(su, "Review Stars")) if su else None
        review_count = to_float(get_ci(keepa, "Reviews: Rating Count")) if keepa else to_float(get_ci(su, "Review Count")) if su else None
        total_offers = to_float(get_ci(keepa, "Total Offer Count")) if keepa else to_float(get_ci(su, "Offers")) if su else None
        category = get_ci(keepa, "Categories: Root") if keepa else get_ci(su, "Category") if su else None
        amazon_url = get_ci(keepa, "URL: Amazon") if keepa else (f"https://www.amazon.com/dp/{asin}" if asin else None)

        master_rows.append(
            {
                "product_name": v.get("product_name"),
                "variation": v.get("variation"),
                "vitacost_price": cost,
                "gtin": v.get("gtin"),
                "sku": v.get("sku"),
                "vitacost_contents_qty": vitacost_contents_qty,
                "amazon_contents_qty": amazon_contents_qty,
                "amazon_bundle_qty": amazon_bundle_qty,
                "distinct_asin_count": distinct_asin_count,
                "vitacost_url": v.get("product_url"),
                "asin": asin,
                "brand": brand,
                "title_amazon": title_amazon,
                "cost": cost,
                "sell_price": sell_price,
                "sell_price_source": sell_price_source,
                "referral_fee": referral_fee,
                "fba_fee": fba_fee,
                "profit": profit,
                "roi_pct": roi_pct,
                "margin_pct": margin_pct,
                "sales_rank": sales_rank,
                "review_rating": review_rating,
                "review_count": review_count,
                "total_offers": total_offers,
                "category": category,
                "amazon_url": amazon_url,
                "flags": ", ".join(flags) if flags else "",
            }
        )

    master_rows.sort(key=lambda r: (r["roi_pct"] is None, -(r["roi_pct"] or 0)))
    return master_rows


# Flags that mean the Profit/ROI/Margin numbers aren't a like-for-like
# comparison (wrong quantity, or possibly the wrong ASIN entirely) - these
# rows get pulled out of the main ranked list so they never look like a
# normal opportunity next to genuinely comparable matches.
REVIEW_FLAG_PREFIXES = ("bundle_mismatch", "ambiguous_match", "contents_count_mismatch")


def split_clean_and_review_rows(
    rows: List[Dict[str, Any]]
) -> tuple:
    clean, review = [], []
    for row in rows:
        flags = row.get("flags") or ""
        if any(flags.startswith(p) or f", {p}" in flags for p in REVIEW_FLAG_PREFIXES):
            review.append(row)
        else:
            clean.append(row)
    return clean, review


# --------------------------------------------------------------------------
# Excel export
# --------------------------------------------------------------------------

def build_columns(source_name: str) -> List[tuple]:
    """Column definitions for the Excel export. source_name labels the
    scrape-source-specific columns (e.g. "Vitacost" or "Moida") so the
    report doesn't say "Vitacost SKU" when the data is actually Moida's."""
    return [
        ("product_name", "Product", 42),
        ("variation", "Variation", 14),
        ("gtin", "GTIN/UPC", 16),
        ("sku", f"{source_name} SKU", 12),
        ("vitacost_contents_qty", f"{source_name} Contents Qty", 18),
        ("amazon_contents_qty", "Amazon Contents Qty", 18),
        ("amazon_bundle_qty", "Amazon Bundle Qty", 15),
        ("distinct_asin_count", "# ASINs for GTIN", 15),
        ("asin", "ASIN", 12),
        ("brand", "Brand", 16),
        ("cost", f"Cost ({source_name})", 13),
        ("sell_price", "Sell Price (Amazon)", 16),
        ("sell_price_source", "Sell Price Source", 24),
        ("referral_fee", "Referral Fee", 12),
        ("fba_fee", "FBA Fee", 10),
        ("profit", "Profit", 10),
        ("roi_pct", "ROI %", 9),
        ("margin_pct", "Margin %", 10),
        ("sales_rank", "Sales Rank", 11),
        ("review_rating", "Rating", 8),
        ("review_count", "Review Count", 12),
        ("total_offers", "Offers", 8),
        ("category", "Category", 20),
        ("flags", "Flags", 42),
        ("vitacost_url", f"{source_name} URL", 45),
        ("amazon_url", "Amazon URL", 45),
    ]

CURRENCY_FIELDS = {"cost", "sell_price", "referral_fee", "fba_fee", "profit"}
PERCENT_FIELDS = {"roi_pct", "margin_pct"}

HEADER_FILL = PatternFill(start_color="FF2F5233", end_color="FF2F5233", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
FLAG_FILL = PatternFill(start_color="FFFFF3CD", end_color="FFFFF3CD", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFF8D7DA", end_color="FFF8D7DA", fill_type="solid")


def _write_sheet(ws, rows: List[Dict[str, Any]], columns: List[tuple]) -> None:
    for col_idx, (_, header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _, _) in enumerate(columns, start=1):
            value = row.get(key)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if key in CURRENCY_FIELDS and value is not None:
                cell.number_format = "$#,##0.00"
            elif key in PERCENT_FIELDS and value is not None:
                cell.number_format = "0.0"

        roi = row.get("roi_pct")
        flags = row.get("flags") or ""
        if roi is not None and roi < 0:
            for c_idx in range(1, len(columns) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = NEGATIVE_FILL
        elif flags:
            for c_idx in range(1, len(columns) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = FLAG_FILL


def export_xlsx(
    clean_rows: List[Dict[str, Any]], review_rows: List[Dict[str, Any]], path: Path, source_name: str = "Moida",
) -> None:
    """Writes two sheets: 'Master' holds only rows whose Profit/ROI/Margin
    are a reliable like-for-like comparison; 'Needs Review' holds every row
    flagged bundle_mismatch, ambiguous_match, or contents_count_mismatch -
    kept for manual verification, but never mixed into the ranked list
    where its numbers could be mistaken for a genuine opportunity."""
    columns = build_columns(source_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    master_ws = wb.active
    master_ws.title = "Master"
    _write_sheet(master_ws, clean_rows, columns)

    review_ws = wb.create_sheet("Needs Review")
    _write_sheet(review_ws, review_rows, columns)

    wb.save(path)


# --------------------------------------------------------------------------
# Entry point
# --------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Merge Moida scrape + ScanUnlimited export + Keepa export into one master sheet."
    )
    parser.add_argument("--vitacost", type=str, default=None, help="Path to the scrape CSV (default: newest moidaus_*.csv in output/)")
    parser.add_argument("--scanunlimited", type=str, required=True, help="Path to ScanUnlimited export CSV")
    parser.add_argument("--keepa", type=str, required=True, help="Path to Keepa export CSV")
    parser.add_argument("--output", type=str, default=None, help="Output .xlsx path (default: output/master_<timestamp>.xlsx)")
    parser.add_argument(
        "--source-name", type=str, default="Moida",
        help="Label for the scrape source in the Excel report and console output. "
             "Cost is read from whichever of 'price' (Vitacost) or 'cart_price' (Moida) is present "
             "in the CSV automatically -- this flag only affects labeling, not that detection.",
    )
    args = parser.parse_args()

    vitacost_path = Path(args.vitacost) if args.vitacost else find_latest_vitacost_csv()
    if not vitacost_path or not vitacost_path.exists():
        raise SystemExit(f"Could not find a {args.source_name} scrape CSV. Pass --vitacost explicitly.")

    print(f"{args.source_name}: {vitacost_path}")
    print(f"ScanUnlimited: {args.scanunlimited}")
    print(f"Keepa:         {args.keepa}")

    vitacost_rows = read_csv(vitacost_path)
    scanunlimited_rows = read_csv(Path(args.scanunlimited))
    keepa_rows = read_csv(Path(args.keepa))

    scanunlimited_index = index_by_code(scanunlimited_rows, "EAN")
    keepa_index = index_by_code_multi(
        keepa_rows, "Product Codes: UPC", "Product Codes: EAN", "Product Codes: GTIN"
    )

    codes = [normalize_code(v.get("gtin")) for v in vitacost_rows]
    su_matched_count = sum(1 for c in codes if c and c in scanunlimited_index)
    keepa_matched_count = sum(1 for c in codes if c and c in keepa_index)
    print(
        f"Diagnostics: {len(scanunlimited_rows)} ScanUnlimited rows -> "
        f"{len(scanunlimited_index)} unique EAN codes indexed -> "
        f"{su_matched_count}/{len(codes)} {args.source_name} rows matched"
    )
    print(
        f"Diagnostics: {len(keepa_rows)} Keepa rows -> "
        f"{len(keepa_index)} unique UPC/EAN/GTIN codes indexed -> "
        f"{keepa_matched_count}/{len(codes)} {args.source_name} rows matched"
    )
    unmatched_su_sample = [get_ci(r, "EAN") for r in scanunlimited_rows[:5]]
    print(f"Diagnostics: first 5 raw ScanUnlimited EAN values as read: {unmatched_su_sample}")
    print(f"Diagnostics: first 5 {args.source_name} gtin values: {[v.get('gtin') for v in vitacost_rows[:5]]}")

    master_rows = build_master_rows(vitacost_rows, scanunlimited_index, keepa_index)

    matched = sum(1 for r in master_rows if "no_amazon_match" not in r["flags"])
    price_changed = sum(1 for r in master_rows if "price_changed_since_scan" in r["flags"])
    no_buy_box = sum(1 for r in master_rows if "no_buy_box" in r["flags"])
    ambiguous = sum(1 for r in master_rows if "ambiguous_match" in r["flags"])
    bundle_mismatch = sum(1 for r in master_rows if "bundle_mismatch" in r["flags"])
    contents_mismatch = sum(1 for r in master_rows if "contents_count_mismatch" in r["flags"])
    print(
        f"Total rows: {len(master_rows)} | Matched: {matched} | Price changed since scan: {price_changed} | "
        f"No buy box: {no_buy_box} | Ambiguous multi-ASIN GTIN: {ambiguous} | "
        f"Bundle mismatch: {bundle_mismatch} | Contents count mismatch: {contents_mismatch}"
    )

    clean_rows, review_rows = split_clean_and_review_rows(master_rows)
    print(f"Master sheet: {len(clean_rows)} reliable rows | Needs Review sheet: {len(review_rows)} flagged rows")

    if args.output:
        output_path = Path(args.output)
    else:
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path("output") / f"master_{timestamp}.xlsx"

    export_xlsx(clean_rows, review_rows, output_path, source_name=args.source_name)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
